from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from models import Node


def save_trees_to_excel(roots: list[Node], filename: str = "wb_categories.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)  # удалим стандартный пустой лист

    for root in roots:
        ws = wb.create_sheet(title=root.name + "- кат")
        ws.append(["ID", "Name", "URL", "Level", "Parent ID"])

        def dfs(node: Node):
            is_leaf = not node.children
            level = 99 if is_leaf else node.level
            ws.append([
                node.id,
                node.name,
                node.url or "",
                level,
                node.parent.id if node.parent else ""
            ])
            for child in node.children:
                dfs(child)

        dfs(root)

        # Автоматическая ширина колонок (по макс длине)
        for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=5), 1):
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(max_length + 2, 12)

    wb.save(filename)